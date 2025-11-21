#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Export Lemon8 reviews from MongoDB to Excel for easy viewing
"""

import os
import sys
from pathlib import Path
from pymongo import MongoClient
from dotenv import load_dotenv
import pandas as pd
from datetime import datetime

if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Load environment
project_root = Path(__file__).parent.parent.parent
env_file = project_root / 'database' / 'scripts' / '.env'
if env_file.exists():
    load_dotenv(env_file)

MONGODB_URI = os.getenv('MONGODB_URI')
MONGODB_DB_NAME = os.getenv('MONGODB_DB_NAME', 'INF2006-Database_Systems')
client = MongoClient(MONGODB_URI, serverSelectionTimeoutMS=5000)
db = client[MONGODB_DB_NAME]

print("\n" + "="*70)
print("[LEMON8 EXPORT] Exporting reviews to Excel")
print("="*70)

# Fetch all lemon8 reviews
reviews = list(db['reviews'].find({'source': 'lemon8'}).sort('analyzed_at', -1))

if not reviews:
    print("\n   [No Lemon8 reviews found!]\n")
    client.close()
    sys.exit(1)

print(f"\n   Found {len(reviews)} reviews\n")

# Prepare data for Excel
export_data = []

for review in reviews:
    # Format amenities
    amenities = review.get('amenities_mentioned', {})
    amenities_str = '; '.join([f"{cat}: {', '.join(items)}" for cat, items in amenities.items()])
    
    # Format key points
    key_points_str = '\n'.join(['• ' + point for point in review.get('key_points', [])])
    pros_str = '\n'.join(['✓ ' + pro for pro in review.get('pros', [])])
    cons_str = '\n'.join(['✗ ' + con for con in review.get('cons', [])])
    
    export_data.append({
        'Estate': review.get('estate', ''),
        'Sentiment': review.get('sentiment', '').upper(),
        'Title': review.get('title', '')[:80],
        'Content': review.get('content', '')[:200],
        'Quality Score': review.get('quality_score', 0),
        'Key Points': key_points_str,
        'Pros': pros_str,
        'Cons': cons_str,
        'Amenities': amenities_str,
        'Hashtags': ', '.join(review.get('hashtags', [])),
        'Author': f"@{review.get('account_handle', '')}",
        'Analyzed': review.get('analyzed_at', '').strftime('%Y-%m-%d %H:%M') if review.get('analyzed_at') else ''
    })

# Create DataFrame
df = pd.DataFrame(export_data)

# Sort by estate and sentiment
df = df.sort_values(['Estate', 'Sentiment', 'Quality Score'], ascending=[True, False, False])

# Create Excel file with formatting
output_file = project_root / 'lemon8_reviews_export.xlsx'

print(f"[Exporting] {len(df)} reviews to Excel...\n")

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Reviews', index=False)
    
    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Reviews']
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 15  # Estate
    worksheet.column_dimensions['B'].width = 12  # Sentiment
    worksheet.column_dimensions['C'].width = 25  # Title
    worksheet.column_dimensions['D'].width = 30  # Content
    worksheet.column_dimensions['E'].width = 10  # Quality Score
    worksheet.column_dimensions['F'].width = 35  # Key Points
    worksheet.column_dimensions['G'].width = 35  # Pros
    worksheet.column_dimensions['H'].width = 35  # Cons
    worksheet.column_dimensions['I'].width = 40  # Amenities
    worksheet.column_dimensions['J'].width = 25  # Hashtags
    worksheet.column_dimensions['K'].width = 15  # Author
    worksheet.column_dimensions['L'].width = 15  # Analyzed
    
    # Format header row
    from openpyxl.styles import Font, PatternFill, Alignment
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Color code sentiment
    sentiment_colors = {
        'POSITIVE': 'C6E0B4',
        'NEUTRAL': 'FFE699',
        'NEGATIVE': 'F4B084'
    }
    
    for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=2, max_col=2):
        for cell in row:
            sentiment = cell.value
            if sentiment in sentiment_colors:
                cell.fill = PatternFill(start_color=sentiment_colors[sentiment], 
                                       end_color=sentiment_colors[sentiment], 
                                       fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
    
    # Set row height for header
    worksheet.row_dimensions[1].height = 30

print(f"[Saved] {output_file}")

# Print summary
print(f"\n[SUMMARY]")
print(f"   Total Reviews: {len(df)}")
print(f"\n   By Sentiment:")
for sentiment in ['POSITIVE', 'NEUTRAL', 'NEGATIVE']:
    count = len(df[df['Sentiment'] == sentiment])
    if count > 0:
        print(f"      {sentiment}: {count}")

print(f"\n   By Estate (Top 10):")
estate_counts = df['Estate'].value_counts().head(10)
for estate, count in estate_counts.items():
    print(f"      {estate}: {count}")

print(f"\n[Output File]")
print(f"   {output_file}")
print(f"   Size: {output_file.stat().st_size / 1024:.1f} KB")
print("\n" + "="*70 + "\n")

client.close()
